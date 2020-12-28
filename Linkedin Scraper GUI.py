#!/usr/bin/env python
# coding: utf-8

# In[1]:


#required installs (i.e. pip3 install in terminal): pandas, selenium, bs4, and possibly chromedriver(it may come with selenium)
#Download Chromedriver from: https://chromedriver.chromium.org/downloads
#To see what version to install: Go to chrome --> on top right click three dot icon --> help --> about Google Chrome
#Move the chrome driver to (/usr/local/bin) -- open finder -> Command+Shift+G -> search /usr/local/bin -> move from downloads

from selenium import webdriver
from bs4 import BeautifulSoup as bs
import time
from datetime import datetime
import pandas as pd
import re
import caffeine
import random
import schedule
import gender_guesser.detector as gender
d = gender.Detector()
import collections
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
get_ipython().run_line_magic('matplotlib', 'inline')
caffeine.on(display=True)
import tkinter as tk
import threading


# In[2]:


root = tk.Tk()

#Defining our variables
tk_page = tk.StringVar()
tk_username = tk.StringVar()
tk_password = tk.StringVar()


canvas = tk.Canvas(root, height = 300 , width = 400, bg ='#49694b')
canvas.pack()

#making all of the frames
frame = tk.Frame(root, bg ='#49694b', bd =6)
frame.place(relx=0.5, rely=0.1, relwidth=0.8, relheight=0.15, anchor='n')

frame2 = tk.Frame(root, bg ='#49694b', bd =6)
frame2.place(relx=0.5, rely=0.25, relwidth=0.8, relheight=0.15, anchor='n')

frame3 = tk.Frame(root, bg ='#49694b', bd =6)
frame3.place(relx=0.5, rely=0.4, relwidth=0.8, relheight=0.15, anchor='n')

button_frame = tk.Frame(root, bg ='#49694b', bd =8)
button_frame.place(relx=0.5, rely=0.6, relwidth=0.8, relheight=0.15, anchor='n')


#Making all of our labels and inputs
label1 = tk.Label(frame, text="Linkedin URL: ", bg='#49694b')
label1.place(relwidth=0.35,relheight=1)

entry1 = tk.Entry(frame)
entry1.place(relx=0.375, relwidth=0.6,relheight=1)


def get_username():
    label2 = tk.Label(frame2, text="Username: ", bg='#49694b')
    label2.place(relwidth=0.35,relheight=1)

    entry2 = tk.Entry(frame2)
    entry2.place(relx=0.375, relwidth=0.6,relheight=1)
    return entry2


def get_password():
    label3 = tk.Label(frame3, text="Password: ", bg='#49694b')
    label3.place(relwidth=0.35,relheight=1)

    entry3 = tk.Entry(frame3)
    entry3.place(relx=0.375, relwidth=0.6, relheight=1)
    return entry3


#Functions to get variables from inputs
def submit1(entry1):
    tk_page.set(entry1)


def submit2(entry2,entry3):
    tk_username.set(entry2)
    tk_password.set(entry3)
    

#Button that calls function and waits until variables are defined to pass
button = tk.Button(button_frame, text= "Check for Existing Project", command=lambda: submit1(entry1.get()))
button.place(relx=0.3, relwidth=0.6,relheight=1)
button.wait_variable(tk_page)

page = tk_page.get()
company_name = page[33:-1]

#See if credetial file exists and create a project if not
try:
    f= open("{}_credentials.txt".format(company_name),"r")
    contents = f.read()
    username = contents.replace("=",",").split(",")[1]
    password = contents.replace("=",",").split(",")[3]
    page = contents.replace("=",",").split(",")[5]
    company_name = page[33:-1]
    post_index = int(contents.replace("=",",").split(",")[7])
    user_index = int(contents.replace("=",",").split(",")[9])
    
    wait = tk.StringVar()
    button = tk.Button(button_frame, text= "Let's Scrape", command=lambda: wait.set("go"))
    button.place(relx=0.3, relwidth=0.6,relheight=1)
    button.wait_variable(wait)
    
except:
    f= open("{}_credentials.txt".format(company_name),"w+")
    entry2 = get_username()
    entry3 = get_password()

    #Button that calls function and waits until variables are defined to pass
    button = tk.Button(button_frame, text= "Let's Srape", command=lambda: submit2(entry2.get(),entry3.get()))
    button.place(relx=0.3, relwidth=0.6,relheight=1)
    button.wait_variable(tk_password)

    #transforming our tk variables into python variables
    username = tk_username.get()
    password = tk_password.get()
    post_index = 1
    user_index = 1
    f.write("username={}, password={}, page={}, post_index={}, user_index={}".format(username,password,page,post_index,user_index))
    f.close()


# In[3]:


#accessing Chromedriver
browser = webdriver.Chrome('chromedriver')

#Open login page
browser.get('https://www.linkedin.com/login?fromSignIn=true&trk=guest_homepage-basic_nav-header-signin')

#Enter login info:
elementID = browser.find_element_by_id('username')
elementID.send_keys(username)

elementID = browser.find_element_by_id('password')
elementID.send_keys(password)
elementID.submit()


# In[4]:


#Scrolls the main page
def scroll():
    #Simulate scrolling to capture all posts
    SCROLL_PAUSE_TIME = 1.5

    # Get scroll height
    last_height = browser.execute_script("return document.body.scrollHeight")

    while True:
        # Scroll down to bottom
        browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")

        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)

        # Calculate new scroll height and compare with last scroll height
        new_height = browser.execute_script("return document.body.scrollHeight")
        if new_height == last_height:
            break
        last_height = new_height


# In[5]:


def scrape_posts(containers):
    
    for container in containers:

        try:
            posted_date = container.find("span",{"class":"visually-hidden"})
            text_box = container.find("div",{"class":"feed-shared-update-v2__description-wrapper ember-view"})
            text = text_box.find("span",{"dir":"ltr"})
            new_likes = container.findAll("li", {"class":"social-details-social-counts__reactions social-details-social-counts__item"})
            new_comments = container.findAll("li", {"class": "social-details-social-counts__comments social-details-social-counts__item"})


            post_dates.append(posted_date.text.strip())
            post_texts.append(text.text.strip())



            try:
                video_box = container.findAll("div",{"class": "feed-shared-update-v2__content feed-shared-linkedin-video ember-view"})
                video_link = video_box[0].find("video", {"class":"vjs-tech"})
                media_links.append(video_link['src'])
                media_type.append("Video")
            except:
                try:
                    image_box = container.findAll("div",{"class": "feed-shared-image__container"})
                    image_link = image_box[0].find("img", {"class":"ivm-view-attr__img--centered feed-shared-image__image feed-shared-image__image--constrained lazy-image ember-view"})
                    media_links.append(image_link['src'])
                    media_type.append("Image")
                except:
                    try:
                        #mutiple shared images
                        image_box = container.findAll("div",{"class": "feed-shared-image__container"})
                        image_link = image_box[0].find("img", {"class":"ivm-view-attr__img--centered feed-shared-image__image lazy-image ember-view"})
                        media_links.append(image_link['src'])
                        media_type.append("Multiple Images")
                    except:
                        try:
                            article_box = container.findAll("div",{"class": "feed-shared-article__description-container"})
                            article_link = article_box[0].find('a', href=True)
                            media_links.append(article_link['href'])
                            media_type.append("Article")
                        except:
                            try:
                                video_box = container.findAll("div",{"class": "feed-shared-external-video__meta"})          
                                video_link = video_box[0].find('a', href=True)
                                media_links.append(video_link['href'])
                                media_type.append("Youtube Video")   
                            except:
                                try:
                                    poll_box = container.findAll("div",{"class": "feed-shared-update-v2__content overflow-hidden feed-shared-poll ember-view"})
                                    media_links.append("None")
                                    media_type.append("Other: Poll, Shared Post, etc")
                                except:
                                    media_links.append("None")
                                    media_type.append("Unknown")



            #Getting Video Views. (The folling three lines prevents class name overlap)
            view_container2 = set(container.findAll("li", {'class':["social-details-social-counts__item"]}))
            view_container1 = set(container.findAll("li", {'class':["social-details-social-counts__reactions","social-details-social-counts__comments social-details-social-counts__item"]}))
            result = view_container2 - view_container1

            view_container = []
            for i in result:
                view_container += i

            try:
                video_views.append(view_container[1].text.strip().replace(' Views',''))

            except:
                video_views.append('N/A')


            try:
                post_likes.append(new_likes[0].text.strip())
            except:
                post_likes.append(0)
                pass

            try:
                post_comments.append(new_comments[0].text.strip())                           
            except:                                                           
                post_comments.append(0)
                pass

        except:
            pass


# In[6]:


def export_post_data():
    
    comment_count = []
    for i in post_comments:
        s = str(i).replace('Comment','').replace('s','').replace(' ','')
        comment_count += [s]
    
    
    data = {
    "Date Posted": post_dates,
    "Media Type": media_type,
    "Post Text": post_texts,
    "Post Likes": post_likes,
    "Post Comments": comment_count,
    "Video Views": video_views,
    "Media Links": media_links
    }


    df = pd.DataFrame(data)

    writer = pd.ExcelWriter("{}_page_posts.xlsx".format(company_name), engine='xlsxwriter')
    df.to_excel(writer, index =False)
    writer.save()
    


# In[7]:


try:
    wb = load_workbook("{}_page_posts.xlsx".format(company_name))
except: 
    browser.get(page + 'posts/')
    time.sleep(2)
    
    #scroll through the page
    scroll()
    
    #get html from page
    company_page = browser.page_source
    linkedin_soup = bs(company_page.encode("utf-8"), "html")
    containers = linkedin_soup.findAll("div",{"class":"occludable-update ember-view"})
    
    #define the variables we want
    post_dates = []
    post_texts = []
    post_likes = []
    post_comments = []
    video_views = []
    media_links = []
    media_type = []
    
    #scrape the post data
    scrape_posts(containers)
    
    #export the df as excel file
    export_post_data()
    


# In[8]:


#Get any saved progress or create new variables
try:
    scraped = pd.read_csv("{}_linkedin_backup.csv".format(company_name))
    liker_names = list(scraped["Id"])
    user_gender = list(scraped["Gender"])
    liker_locations = list(scraped["Location"])
    liker_headlines = list(scraped["Headline"])
    user_bios = list(scraped["Bio"])
    est_ages = list(scraped["Age"])
    influencers = list(scraped["Followed Influencers"])
    companies = list(scraped["Followed Companies"])
except:
    liker_names = []
    user_gender = []
    liker_locations = []
    liker_headlines = []
    user_bios = []
    est_ages = []
    influencers = []
    companies = []
    pass


# In[9]:


#Scrolls popups
def scroll_popup(class_name):
    #Simulate scrolling to capture all posts
    SCROLL_PAUSE_TIME = 1.5

    # Get scroll height
    js_code = "return document.getElementsByClassName('{}')[0].scrollHeight".format(class_name)
    last_height = browser.execute_script(js_code)

    while True:
        # Scroll down to bottom
        path = "//div[@class='{}']".format(class_name)
        browser.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", browser.find_element_by_xpath(path))

        # Wait to load page
        time.sleep(SCROLL_PAUSE_TIME)

        # Calculate new scroll height and compare with last scroll height
        new_height = browser.execute_script(js_code)
        if new_height == last_height:
            break
        last_height = new_height
        
        


# In[10]:


#Function that estimates user age based on earliest school date or earlier work date
def est_age():

    browser.switch_to.window(browser.window_handles[1])
    date = datetime.today()
    current_year = date.strftime("%Y")
    school_start_year = "9999"
    work_start_year = "9999"

    #Get page source
    user_profile = browser.page_source
    user_profile = bs(user_profile.encode("utf-8"), "html")


    #Look for earliest university start date
    try:
        grad_year = user_profile.findAll('p',{"class":"pv-entity__dates t-14 t-black--light t-normal"})
        
        if grad_year == []:
            browser.execute_script("window.scrollTo(0, 1000);")
            user_profile = browser.page_source
            user_profile = bs(user_profile.encode("utf-8"), "html")
            grad_year = user_profile.findAll('p',{"class":"pv-entity__dates t-14 t-black--light t-normal"})
            
        
        for d in grad_year:
            year = d.find('time').text.strip().replace(' ', '')
            start_year = re.sub(r'[a-zA-Z]', r'', year)
            start_year = start_year[0:4]
            if start_year < school_start_year:
                        school_start_year = start_year
    except:
        pass
    

    #Look for earlies work date
    try:
        #Click see more if it's there
        try:
            browser.find_element_by_xpath("//button[@class='pv-profile-section__see-more-inline pv-profile-section__text-truncate-toggle link-without-visited-state']").click()
        except:
            time.sleep(1)
            pass

        work_start = user_profile.findAll('h4', {"class":"pv-entity__date-range t-14 t-black--light t-normal"})


        for d in work_start:
            start_date = d.find('span',class_=None)
            start_date = start_date.text.strip().replace(' ', '')
            start_date = re.sub(r'[a-zA-Z]', r'', start_date)
            start_year = start_date[0:4]
            if start_year < work_start_year:
                    work_start_year = start_year
    except:
        pass

    # Compare work and school start dates to avoid adult degress
    if school_start_year < work_start_year:
        #Estimate age based on avg university start age of 18
        est_birth_year = int(school_start_year) - 18
        est_age = int(current_year) - est_birth_year

    else:
        #Estimate age based on avg post college work start date of 22
        est_birth_year = int(work_start_year) - 22
        est_age = int(current_year) - est_birth_year

    if est_age == -7961 or est_age == -7957:
        est_age = 'unknown'
    
    return est_age
        


# In[11]:


#Function that Scrapes user data
def get_user_data():
    
    global skip_count
      
    user_profile = browser.page_source
    user_profile = bs(user_profile.encode("utf-8"), "html")

    name = user_profile.find('li',{'class':"inline t-24 t-black t-normal break-words"})
    name = name.text.strip()

    #Make sure liker isn't a duplicate
    if name not in liker_names:

        skip_count = 0
        liker_names.append(name)
        split_name = name.split(" ", 2)
        #Get Liker Gender
        user_gender.append(d.get_gender(split_name[0])+"^ ")

        try:
            #Get Liker Location
            location = user_profile.find('li',{'class':"t-16 t-black t-normal inline-block"})
            liker_locations.append(location.text.strip()+"^ ")
        except:
            liker_locations.append("No Location")

        try:
            #Get Liker Headline
            headline = user_profile.find('h2',{"class":"mt1 t-18 t-black t-normal break-words"})
            liker_headlines.append(headline.text.strip())
        except:
            liker_headlines.append("No Headline")


        #Get Liker Bio
        try:
            browser.find_element_by_xpath("//a[@id='line-clamp-show-more-button']").click()
            time.sleep(1)
            user_profile = browser.page_source
            user_profile = bs(user_profile.encode("utf-8"), "html")
            bio = user_profile.findAll("span",{"class":"lt-line-clamp__raw-line"})
            user_bios.append(bio[0].text.strip())
        except:
            try:
                bio_lines = []
                bios = user_profile.findAll('span',{"class":"lt-line-clamp__line"})
                for b in bios:
                    bio_lines.append(b.text.strip())
                bio = ",".join(bio_lines).replace(",", ". ")
                user_bios.append(bio)

            except:
                user_bios.append('No Bio')
                pass

        #Get estimated age using our age function
        age = est_age()
        est_ages.append(age)



        #Click see more on user interests
        try: 
            interest_path = "//a[@data-control-name='view_interest_details']"
            browser.find_element_by_xpath(interest_path).click()
        except:
            scroll()
            time.sleep(1)
            try:
                interest_path = "//a[@data-control-name='view_interest_details']"
                browser.find_element_by_xpath(interest_path).click()
            except:
                influencers.append("No Influencers^ ")
                companies.append("No Companies^ ")
                return

        time.sleep(1)

        #Scrape the influencers the user follows
        try:
            influencer_path = "//a[@id='pv-interests-modal__following-influencers']"
            browser.find_element_by_xpath(influencer_path).click()

            #Scroll the end of list
            class_name = 'entity-all pv-interests-list ml4 pt2 ember-view'
            #interest_box_path = "//div[@class='entity-all pv-interests-list ml4 pt2 ember-view']"
            scroll_popup(class_name)

            influencer_page = browser.page_source
            influencer_page = bs(influencer_page.encode("utf-8"), "html")
            influencer_list = influencer_page.findAll("li",{"class":"entity-list-item"})


            user_influencers = ""
            for i in influencer_list:
                name = i.find("span",{"class":"pv-entity__summary-title-text"})
                user_influencers += name.text.strip() + "^ "

            influencers.append(user_influencers)


        except:
            influencers.append("No Influencers^ ")



        #Scrape the companies the user follows
        try:
            company_path = "//a[@id='pv-interests-modal__following-companies']"
            browser.find_element_by_xpath(company_path).click()

            time.sleep(2)

            #Scroll the end of list
            class_name = 'entity-all pv-interests-list ml4 pt2 ember-view'
            #interest_box_path = "//div[@class='entity-all pv-interests-list ml4 pt2 ember-view']"
            scroll_popup(class_name)


            company_page = browser.page_source
            company_page = bs(company_page.encode("utf-8"), "html")
            company_list = company_page.findAll("li",{"class":"entity-list-item"})


            user_companies = ""
            for i in company_list:
                name = i.find("span",{"class":"pv-entity__summary-title-text"})
                user_companies += name.text.strip() + "^ "

            companies.append(user_companies)

        except:
            companies.append("No Companies^ ")

    else:
        skip_count+=1
        time.sleep(random.randint(2,7))
        


# In[12]:


def word_counter(words):
    wordcount = {}
    for word in words.split('^ '):
        word = word.replace("\"","")
        word = word.replace("!","")
        word = word.replace("â€œ","")
        word = word.replace("â€˜","")
        word = word.replace("*","")
        word = word.replace("?","")
        word = word.replace("mostly_male","male")
        word = word.replace("mostly_female","female")
        
        if word != "No Influencers" and word != "No Companies" and word != "unknown" and word != "":
            if word not in wordcount:
                wordcount[word] = 1
            else:
                wordcount[word] += 1
        else:
            pass
            
    return wordcount


# In[13]:


def get_df(wc):
    
    total_scraped = len(user_gender)
    
    trimmed_count = collections.Counter(wc).most_common(200)

    words = []
    count = []
    percent = []
    for item in trimmed_count:
        words.append(item[0])
        count.append(item[1])
        
    for c in count:
        percent.append(round(((c/total_scraped) * 100), 2))
        

    data = {"Word": words,"Count": count, "Percentage": percent}

    df = pd.DataFrame(data, index =None)
    return df


# In[14]:


def clean_list(interest):
    clean_list = []
    for item in interest:
        clean = item.replace('^','')
        clean_list.append(clean.title())
    return clean_list


# In[15]:


def clean_interests(interest):
    clean_list = []
    for item in interest:
        clean = item.replace('^',',')
        clean_list.append(clean)
    return clean_list


# In[16]:


def count_interests():
    company_list = ",".join(companies).replace(',','')
    company_count = word_counter(company_list)
    common_companies = get_df(company_count)

    influencer_list = ",".join(influencers).replace(',','')
    influencer_count = word_counter(influencer_list)
    common_influencers = get_df(influencer_count)
    
    gender_list = ",".join(user_gender).replace(',','')
    gender_count = word_counter(gender_list)
    common_genders = get_df(gender_count)

    location_list = ",".join(liker_locations).replace(',','')
    location_count = word_counter(location_list)
    common_locations = get_df(location_count)
    
    return common_companies, common_influencers, common_genders, common_locations


# In[27]:


def plot_interests(df1,df2,df3,df4):
    company_plot = df1[0:24].plot.barh(x='Word',y='Percentage')
    company_plot.invert_yaxis()
    company_plot.set_ylabel('Companies')
    company_plot.figure.savefig("c_plot.png", dpi = 100, bbox_inches = "tight")

    influencer_plot = df2[0:24].plot.barh(x='Word',y='Percentage')
    influencer_plot.invert_yaxis()
    influencer_plot.set_ylabel('Influencers')
    influencer_plot.figure.savefig("i_plot.png", dpi = 100, bbox_inches = "tight")
    
    gender_plot = df3[0:24].plot.barh(x='Word',y='Percentage')
    gender_plot.invert_yaxis()
    gender_plot.set_ylabel('Gender')
    gender_plot.figure.savefig("g_plot.png", dpi = 100, bbox_inches = "tight")

    location_plot = df4[0:24].plot.barh(x='Word',y='Percentage')
    location_plot.invert_yaxis()
    location_plot.set_ylabel('Locations')
    location_plot.figure.savefig("l_plot.png", dpi = 100, bbox_inches = "tight")
    
    plt.close('all')


# In[18]:


def export_df():
    #Constructing Pandas Dataframe
    data = {
        "Gender": clean_list(user_gender),
        "Location": clean_list(liker_locations),
        "Age": est_ages,
        "Headline": liker_headlines,
        "Bio": user_bios,
        "Followed Influencers": clean_interests(influencers),
        "Followed Companies": clean_interests(companies)
    }

    df = pd.DataFrame(data)
    
    #Make backup data from to save our progress
    backup_data = {
        "Id": liker_names,
        "Gender": user_gender,
        "Location": liker_locations,
        "Age": est_ages,
        "Headline": liker_headlines,
        "Bio": user_bios,
        "Followed Influencers": influencers,
        "Followed Companies": companies    
    }
    
    backup_df = pd.DataFrame(backup_data)
    
    
    #Make a df of ages stats
    age_list = []
    for a in df["Age"]:
        if a != "unknown":
            age_list.append(int(a))
        else:
            pass
        
    age_data = {"Ages": age_list}    
    
    ages = pd.DataFrame(age_data)
    age_stats = ages.describe()
    age_stats = pd.DataFrame(age_stats)
    

    #Exporting csv to program folder for backup
    backup_df.to_csv("{}_linkedin_backup.csv".format(company_name), encoding='utf-8', index=False)
    
    #Get data frames of interest counts
    common_companies, common_influencers, common_genders, common_locations = count_interests()
    
    #Plot the interest counts
    plot_interests(common_companies, common_influencers, common_genders, common_locations)
    
    time.sleep(1)
    
    #Create/Update Excel file
    writer = pd.ExcelWriter("{}_linkedin.xlsx".format(company_name), engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Page Egagers', index=False)
    common_companies.to_excel(writer, sheet_name='Company Interest', index=False)
    common_influencers.to_excel(writer, sheet_name='Influencer Interest', index=False)
    age_stats.to_excel(writer, sheet_name='Demographic Stats', index=True)
    writer.save()
    
    wb = load_workbook("{}_linkedin.xlsx".format(company_name))

    #Adding plots to the sheets
    cws = wb["Company Interest"]
    c_img = openpyxl.drawing.image.Image('c_plot.png')
    c_img.anchor = 'E5'
    cws.add_image(c_img)

    iws = wb["Influencer Interest"]
    i_img = openpyxl.drawing.image.Image('i_plot.png')
    i_img.anchor = 'E5'
    iws.add_image(i_img)
    
    dws = wb["Demographic Stats"]
    g_img = openpyxl.drawing.image.Image('g_plot.png')
    g_img.anchor = 'D2'
    dws.add_image(g_img)
    l_img = openpyxl.drawing.image.Image('l_plot.png')
    l_img.anchor = 'B21'
    dws.add_image(l_img)

    #Save Excel file
    wb.save("{}_linkedin.xlsx".format(company_name))
    
    #Keep Track of where we are in the foller list
    f= open("{}_credentials.txt".format(company_name),"w+")
    f.write("username={}, password={}, page={}, post_index={}, user_index={}".format(username,password,page,post_index,user_index))
    f.close()


# In[19]:


def current_time():
    current_time = datetime.now().strftime("%H:%M")
    return current_time

#Keeping track of number of page visits per day to stay under the limit
daily_count = 0
daily_limit = 200

#The path of the block that we need to select to scroll
block_path = "//div[@class='artdeco-modal__content social-details-reactors-modal__content ember-view']"


# In[20]:


#Scraping the list of likers from the post
def scrape_post_likers(): 
    
    #Global variable we will reference/iterate
    global daily_count
    global daily_limit
    global block_path
    global post_index
    skip_count = 0
    
    #Liker link number we will iterate to the path
    global user_index
    
    #Get Length of Entire Liker List
    class_name = 'artdeco-modal__content social-details-reactors-modal__content ember-view'
    js_code = "return document.getElementsByClassName('{}')[0].scrollHeight".format(class_name)
    last_height = browser.execute_script(js_code)
    
    
    while True:
        
        if skip_count <=10:
            pass
        else:
            print("Looks like we've scaped this post already. Let's go to the next.")
            break
            
        time.sleep(random.randint(3,15))
        
        while True:

            # Switch to the new window and scroll and retry if it wasn't found
            try:
                path = "//ul[@class='artdeco-list artdeco-list--offset-1']/li[{}]".format(user_index) 
                user_page = browser.find_element_by_xpath(path)
                user_page.click()
                time.sleep(random.randint(1,3))
                browser.switch_to.window(browser.window_handles[1])
                break
            except:
                print("One sec, I need to scroll to the next liker")
                path = "//div[@class='{}']".format(class_name)
                browser.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight", browser.find_element_by_xpath(path))

                time.sleep(2)
                
                # Calculate new scroll height and compare with last scroll height
                new_height = browser.execute_script(js_code)
                if new_height == last_height:
                    print("All users on post {} have been scraped".format(post_index))
                    break
                else:
                    last_height = new_height
        

        time.sleep(random.randint(2,5))

        try:
            #Scrape the users page with function
            get_user_data()
            browser.close()
            time.sleep(2)
            # Switch back to the first tab with URL A
            browser.switch_to.window(browser.window_handles[0])
            
        except:
            print("Let's export")
            export_df()
            post_index+=1
            user_index = 1
            time.sleep(2)
            break

                
                
        #Iterate save progress if multiple of 10
        user_index+=1
        daily_count+=1

        if user_index % 10 == 0:
            try:
                export_df()
                print(user_index)
            except:
                print("Hmmm...Failed to Export.")


            #Random long sleep function to prevent linkedin rate limit
            time.sleep(random.randint(200,1200))

            #Stop if reached daily page view limit
            if daily_count >= daily_limit:
                print("Daily page limit of "+daily_limit+" has been reached. Stopping for the day to prevent auto signout.")
                while current_time() > "00:01":
                    schedule.run_pending()
                    time.sleep(60)
                daily_count = 0

            #Stop for the night
            while current_time() < "07:05":
                schedule.run_pending()
                time.sleep(60)

        else:
            time.sleep(1)


# In[21]:


#Advanced scrolling
#Open the list of likers for each post and get any new users
def get_next_post():

    global post_index
    
    browser.get(page + 'posts/')
    time.sleep(2)
    
    last_height = browser.execute_script("return document.body.scrollHeight")
    
    while True:
    
        try:
            likers = browser.find_element_by_xpath("(//ul[@class='social-details-social-counts ember-view'])[{}]/li".format(post_index))
            likers.click()
            time.sleep(2)
            scrape_post_likers()
            browser.switch_to.window(browser.window_handles[0])
            browser.get(page + 'posts/')
            time.sleep(2)
            
        except:
            print("One sec..I need to scroll to the next post.")
            browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            
            # Calculate new scroll height and compare with last scroll height
            new_height = browser.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                print("All posts have been scraped")
                break
            else:
                last_height = new_height


# In[22]:


#Calling the Master function
get_next_post()
root.mainloop()


# In[ ]:




